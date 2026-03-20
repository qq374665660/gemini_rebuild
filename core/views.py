from django.shortcuts import render, redirect, get_object_or_404
from django.conf import settings
from django.urls import reverse
from django.db.models import Q, Count
from django.http import HttpResponse, Http404, FileResponse, JsonResponse
from django.utils.encoding import force_str
import shutil
import os
import re
import mimetypes
import time
import tempfile
import subprocess
from difflib import SequenceMatcher
from decimal import Decimal
from django.db import transaction
from .models import Project, ProjectAnalysis, APIConfig, MetricsItem, ExpenseImport, ExpenseSnapshot, ExpenseMapping
from .forms import ProjectForm
import openpyxl
from django.contrib import messages
import json
import requests
from django.utils import timezone
from pathlib import Path
from datetime import datetime, timedelta
from io import BytesIO
from urllib.parse import quote
from django.views.decorators.http import require_POST
from .docx_task_extractor import extract_task_docx
from django.utils.dateparse import parse_date
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

def _normalize_path(path_str):
    normalized = os.path.normpath(path_str)
    if os.name == 'nt':
        normalized = os.path.normcase(normalized)
    return normalized

def _resolve_for_compare(path_str):
    try:
        resolved = Path(path_str).expanduser().resolve(strict=False)
    except Exception:
        resolved = Path(os.path.normpath(path_str))
    return _normalize_path(str(resolved))

def _is_within_root(target_path, root_path):
    try:
        target_norm = _resolve_for_compare(target_path)
        root_norm = _resolve_for_compare(root_path)
        return os.path.commonpath([target_norm, root_norm]) == root_norm
    except Exception:
        return False

def _build_abs_path(base_dir, maybe_relative):
    if not maybe_relative:
        return str(base_dir)
    maybe_relative = str(maybe_relative)
    if os.path.isabs(maybe_relative):
        return maybe_relative
    return os.path.join(str(base_dir), maybe_relative)

def _relpath_for_tree(item_path, base_path):
    try:
        rel_path = os.path.relpath(item_path, base_path)
    except ValueError:
        rel_path = str(item_path)
    return rel_path.replace('\\', '/')

def _compute_progress_node(project, today):
    completed_statuses = {'未立项', '结题', '终止'}
    completed = project.status in completed_statuses or bool(project.actual_completion_date)

    start_date = project.start_date
    end_date = project.extension_date or project.planned_end_date
    progress_available = bool(start_date and end_date)

    midpoint_date = None
    total_days = None
    elapsed_days = None
    progress_percent = 0
    days_remaining = None
    overdue_days = None
    midterm_reached = False
    overdue = False

    if progress_available:
        total_days = max((end_date - start_date).days, 1)
        elapsed_days = (today - start_date).days
        progress_ratio = elapsed_days / total_days
        progress_ratio = min(max(progress_ratio, 0), 1)
        progress_percent = int(round(progress_ratio * 100))
        midpoint_date = start_date + timedelta(days=total_days // 2)
        midterm_reached = today >= midpoint_date and today >= start_date
        days_remaining = (end_date - today).days
        overdue = today > end_date and not completed
        if overdue:
            overdue_days = (today - end_date).days

    if completed:
        node_status = '已完成'
        status_key = 'complete'
    elif not progress_available:
        node_status = '缺少日期'
        status_key = 'unknown'
    elif today < start_date:
        node_status = '未开始'
        status_key = 'pending'
    elif overdue:
        node_status = '超期'
        status_key = 'overdue'
    elif midterm_reached:
        node_status = '已到中期'
        status_key = 'midterm'
    else:
        node_status = '未到中期'
        status_key = 'ontrack'

    return {
        'start_date': start_date,
        'end_date': end_date,
        'midpoint_date': midpoint_date,
        'progress_available': progress_available,
        'progress_percent': progress_percent,
        'total_days': total_days,
        'elapsed_days': elapsed_days,
        'days_remaining': days_remaining,
        'overdue_days': overdue_days,
        'midterm_reached': midterm_reached,
        'overdue': overdue,
        'node_status': node_status,
        'status_key': status_key,
        'midpoint_percent': 50 if progress_available else None,
    }


def _clean_match_text(value):
    if value is None:
        return ''
    text_value = str(value).strip().lower()
    if not text_value:
        return ''
    text_value = re.sub(r'\s+', '', text_value)
    text_value = re.sub(r'[^0-9a-zA-Z\u4e00-\u9fff]+', '', text_value)
    return text_value


def _similarity_score(left, right):
    if not left or not right:
        return 0.0
    if left in right or right in left:
        return 1.0
    return SequenceMatcher(None, left, right).ratio()


def _parse_threshold(value, default=0.85):
    try:
        threshold = float(value)
    except (TypeError, ValueError):
        threshold = default
    if threshold < 0.5:
        threshold = 0.5
    if threshold > 0.98:
        threshold = 0.98
    return threshold


def _split_query_terms(query):
    if not query:
        return []
    terms = re.split(r'[,，;；\s]+', str(query).strip())
    return [term for term in terms if term]


def _parse_decimal_param(value):
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    match = re.search(r'-?\d+(?:\.\d+)?', text)
    if not match:
        return None
    try:
        return Decimal(match.group(0))
    except Exception:
        return None


def _extract_list_param(request, key):
    values = []
    for item in request.GET.getlist(key):
        if item is None:
            continue
        text = str(item).strip()
        if text:
            values.append(text)
    return values


def _apply_multi_value_filter(queryset, field_name, values):
    if not values:
        return queryset
    include_blank = '__blank__' in values
    clean_values = [v for v in values if v != '__blank__']
    if include_blank and clean_values:
        return queryset.filter(
            Q(**{f'{field_name}__in': clean_values}) |
            Q(**{f'{field_name}__isnull': True}) |
            Q(**{f'{field_name}__exact': ''})
        )
    if include_blank:
        return queryset.filter(
            Q(**{f'{field_name}__isnull': True}) |
            Q(**{f'{field_name}__exact': ''})
        )
    return queryset.filter(**{f'{field_name}__in': clean_values})


def _apply_project_filters(queryset, request):
    query = request.GET.get('q', '').strip()
    terms = _split_query_terms(query)
    for term in terms:
        queryset = queryset.filter(
            Q(name__icontains=term) |
            Q(project_id__icontains=term) |
            Q(project_lead__icontains=term) |
            Q(contact_person__icontains=term) |
            Q(ownership__icontains=term) |
            Q(managing_unit__icontains=term) |
            Q(level__icontains=term) |
            Q(project_type__icontains=term) |
            Q(role__icontains=term) |
            Q(status__icontains=term) |
            Q(research_content__icontains=term) |
            Q(remarks__icontains=term)
        )

    year_values = _extract_list_param(request, 'year')
    if year_values:
        year_numbers = []
        for value in year_values:
            try:
                year_numbers.append(int(value))
            except (TypeError, ValueError):
                continue
        if year_numbers:
            queryset = queryset.filter(start_year__in=year_numbers)

    status_values = _extract_list_param(request, 'status')
    queryset = _apply_multi_value_filter(queryset, 'status', status_values)

    level_values = _extract_list_param(request, 'level')
    queryset = _apply_multi_value_filter(queryset, 'level', level_values)

    ownership_values = _extract_list_param(request, 'ownership')
    queryset = _apply_multi_value_filter(queryset, 'ownership', ownership_values)

    type_values = _extract_list_param(request, 'project_type')
    queryset = _apply_multi_value_filter(queryset, 'project_type', type_values)

    role_values = _extract_list_param(request, 'role')
    queryset = _apply_multi_value_filter(queryset, 'role', role_values)

    unit_values = _extract_list_param(request, 'managing_unit')
    queryset = _apply_multi_value_filter(queryset, 'managing_unit', unit_values)

    lead_values = _extract_list_param(request, 'project_lead')
    queryset = _apply_multi_value_filter(queryset, 'project_lead', lead_values)

    min_budget = _parse_decimal_param(request.GET.get('min_budget'))
    if min_budget is not None:
        queryset = queryset.filter(total_budget__gte=min_budget)

    max_budget = _parse_decimal_param(request.GET.get('max_budget'))
    if max_budget is not None:
        queryset = queryset.filter(total_budget__lte=max_budget)

    start_date_from = parse_date(request.GET.get('start_date_from', '').strip())
    if start_date_from:
        queryset = queryset.filter(start_date__gte=start_date_from)

    start_date_to = parse_date(request.GET.get('start_date_to', '').strip())
    if start_date_to:
        queryset = queryset.filter(start_date__lte=start_date_to)

    end_date_from = parse_date(request.GET.get('end_date_from', '').strip())
    if end_date_from:
        queryset = queryset.filter(planned_end_date__gte=end_date_from)

    end_date_to = parse_date(request.GET.get('end_date_to', '').strip())
    if end_date_to:
        queryset = queryset.filter(planned_end_date__lte=end_date_to)

    return queryset


def _get_completion_date(project):
    return project.actual_completion_date or project.extension_date or project.planned_end_date


def _move_path_with_backup(src_path, dest_path):
    if not src_path or not os.path.exists(src_path):
        return False, None
    os.makedirs(os.path.dirname(dest_path), exist_ok=True)
    backup_path = None
    if os.path.exists(dest_path):
        backup_path = dest_path + '_backup_' + str(int(time.time()))
        shutil.move(dest_path, backup_path)
    shutil.move(src_path, dest_path)
    return True, backup_path

def _get_project_folder_name(project):
    sanitized_name = re.sub(r'[\\/*?"<>|]', '_', project.name)
    return f"{project.start_year}-{project.status}-{project.project_id}-{sanitized_name}"

def create_project_directory_structure(project):
    """根据PRD文档4.3节要求创建课题目录结构"""
    folder_name = _get_project_folder_name(project)
    
    # 确保projects根目录存在
    projects_root = str(settings.PROJECTS_ROOT)
    os.makedirs(projects_root, exist_ok=True)
    
    base_dir = os.path.join(projects_root, folder_name)
    
    # 按照文档4.3节规定的目录结构
    required_dirs = [
        '01_申报',
        '02_立项', 
        '03_开题及任务书',
        '04_中期',
        '05_变更',
        '06_结题',
        '07_其它',
    ]
    
    try:
        # 确保项目基础目录存在
        os.makedirs(base_dir, exist_ok=True)

        # 兼容旧目录名称：03_开题 -> 03_开题及任务书
        legacy_dir = os.path.join(base_dir, '03_开题')
        new_dir = os.path.join(base_dir, '03_开题及任务书')
        if os.path.isdir(legacy_dir) and not os.path.exists(new_dir):
            try:
                os.rename(legacy_dir, new_dir)
            except Exception as e:
                print(f"重命名子目录失败: {e}")
        
        # 创建所有必需的子目录
        for d in required_dirs:
            dir_path = os.path.join(base_dir, d)
            os.makedirs(dir_path, exist_ok=True)
        
        # 更新project的directory_path
        if project.directory_path != base_dir:
            project.directory_path = base_dir
            project.save(update_fields=['directory_path'])
            
        print(f"项目目录创建成功: {base_dir}")  # 调试信息
        
    except Exception as e:
        print(f"创建项目目录失败: {e}")  # 调试信息
        # 如果创建失败，至少设置一个基本路径
        if not project.directory_path:
            project.directory_path = base_dir
            project.save(update_fields=['directory_path'])

def rename_project_folder(request, project, old_path):
    """重命名项目文件夹"""
    import time
    new_folder_name = _get_project_folder_name(project)
    projects_root = str(settings.PROJECTS_ROOT)
    new_path = os.path.join(projects_root, new_folder_name)
    
    if old_path != new_path:
        try:
            if os.path.exists(old_path):
                # 如果新路径已存在，先删除或重命名
                if os.path.exists(new_path):
                    backup_path = new_path + '_backup_' + str(int(time.time()))
                    shutil.move(new_path, backup_path)
                    messages.warning(request, f"原目录已备份为: {os.path.basename(backup_path)}")
                
                shutil.move(old_path, new_path)
                messages.success(request, f"项目目录已重命名为: {new_folder_name}")
            
            # 无论是否移动成功，都更新数据库中的路径
            project.directory_path = new_path
            project.save(update_fields=['directory_path'])
            
        except Exception as e:
            messages.error(request, f"重命名项目目录失败: {e}")
            # 即使重命名失败，也要更新数据库路径以保持一致性
            project.directory_path = new_path
            project.save(update_fields=['directory_path'])
    elif not os.path.exists(new_path):
        create_project_directory_structure(project)

def get_directory_tree(path, base_path=None):
    def format_file_size(size_bytes):
        if size_bytes == 0:
            return "0 B"
        size_names = ["B", "KB", "MB", "GB"]
        i = 0
        while size_bytes >= 1024 and i < len(size_names) - 1:
            size_bytes /= 1024.0
            i += 1
        return f"{size_bytes:.1f} {size_names[i]}"
    
    tree = []
    if base_path is None:
        base_path = path
    if not os.path.isdir(path):
        return []
    for item in sorted(os.listdir(path)):
        item_path = os.path.join(path, item)
        rel_path = _relpath_for_tree(item_path, base_path)
        
        node = {
            'name': item,
            'path': rel_path,
            'type': 'folder' if os.path.isdir(item_path) else 'file',
        }
        
        if os.path.isdir(item_path):
            children = get_directory_tree(item_path, base_path)
            node['children'] = children
            # 复用子节点结果，避免为统计信息重复扫描目录
            file_count = sum(1 for child in children if child.get('type') == 'file')
            folder_count = sum(1 for child in children if child.get('type') == 'folder')
            node['file_count'] = file_count
            node['folder_count'] = folder_count
            node['total_items'] = file_count + folder_count
        else:
            # Add file size and modification time for files
            try:
                file_size = os.path.getsize(item_path)
                node['size'] = format_file_size(file_size)
                node['size_bytes'] = file_size
                # 添加修改时间
                import datetime
                mtime = os.path.getmtime(item_path)
                node['modified_time'] = datetime.datetime.fromtimestamp(mtime).strftime('%Y-%m-%d %H:%M')
            except (OSError, IOError):
                node['size'] = "未知"
                node['size_bytes'] = 0
                node['modified_time'] = "未知"
            node['children'] = []
        
        tree.append(node)
    return tree

def create_project_view(request):
    if request.method == 'POST':
        form = ProjectForm(request.POST)
        if form.is_valid():
            project = form.save()
            create_project_directory_structure(project)
            return redirect('project_list')
    else:
        form = ProjectForm()
    return render(request, 'core/create_project.html', {'form': form})

def _normalize_text(value):
    if value is None:
        return ''
    return str(value).strip()

def _pick_first(*values):
    for value in values:
        text = _normalize_text(value)
        if text:
            return text
    return ''

def _safe_date(year, month, day):
    try:
        return datetime(int(year), int(month), int(day)).date()
    except ValueError:
        try:
            return datetime(int(year), int(month), 1).date()
        except ValueError:
            return None

def _find_dates(text):
    if not text:
        return []
    results = []
    patterns = [
        r'(\d{4})[./-](\d{1,2})[./-](\d{1,2})',
        r'(\d{4})年(\d{1,2})月(\d{1,2})日?',
    ]
    for pattern in patterns:
        for match in re.finditer(pattern, text):
            date_obj = _safe_date(match.group(1), match.group(2), match.group(3))
            if date_obj and date_obj not in results:
                results.append(date_obj)
    for match in re.finditer(r'(\d{4})年(\d{1,2})月', text):
        date_obj = _safe_date(match.group(1), match.group(2), 1)
        if date_obj and date_obj not in results:
            results.append(date_obj)
    return results

def _extract_date_range(text):
    dates = _find_dates(text)
    if not dates:
        return None, None
    if len(dates) == 1:
        return dates[0], None
    return dates[0], dates[-1]

def _coerce_number(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = _normalize_text(value)
    if not text:
        return None
    match = re.search(r'-?\d+(?:\.\d+)?', text)
    if match:
        try:
            return float(match.group(0))
        except ValueError:
            return None
    return None

def _format_number(value):
    if value is None:
        return ''
    if isinstance(value, (int,)):
        return str(value)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return f"{value:.2f}".rstrip('0').rstrip('.')

def _map_ownership(text):
    if not text:
        return ''
    if '地下空间' in text:
        return '地下空间'
    if '西勘院' in text:
        return '西勘院'
    return ''

def _is_placeholder_name(text):
    if not text:
        return True
    compact = re.sub(r'\s+', '', text)
    return compact in {'姓名', '姓名:', '姓', '姓名/'} or compact == '姓名'

def _extract_budget_totals(rows):
    if not rows:
        return None
    for keyword in ('经费支出（合计）', '经费来源（合计）'):
        for row in rows:
            if keyword in _normalize_text(row.get('预算科目名称')):
                return row
    return rows[0]

def _parse_funding_text(text):
    if not text:
        return {}
    result = {}
    text = str(text)
    patterns = [
        (r'(?:\u9662\u4e13\u9879)[:\uff1a]?\s*(\d+(?:\.\d+)?)', 'institute_funding'),
        (r'(?<!\u9662)\u4e13\u9879[:\uff1a]?\s*(\d+(?:\.\d+)?)', 'external_funding'),
        (r'\u81ea\u7b79[:\uff1a]?\s*(\d+(?:\.\d+)?)', 'unit_funding'),
        (r'(?:\u5408\u8ba1|\u603b\u8ba1)[:\uff1a]?\s*(\d+(?:\.\d+)?)', 'total_budget'),
    ]
    for pattern, key in patterns:
        match = re.search(pattern, text)
        if match and key not in result:
            try:
                result[key] = float(match.group(1))
            except ValueError:
                continue
    return result


def _map_docx_to_project_fields(extracted):
    basic = extracted.get('basic_info', {}) or {}
    fields = extracted.get('topic_info', {}).get('fields', {}) or {}
    budget_rows = extracted.get('budget_summary', {}).get('rows', []) or []

    project_id = _pick_first(basic.get('\u8bfe\u9898\u7f16\u53f7'), fields.get('\u8bfe\u9898\u7f16\u53f7'))
    name = _pick_first(basic.get('\u8bfe\u9898\u540d\u79f0'), fields.get('\u8bfe\u9898\u540d\u79f0'))
    managing_unit = _pick_first(
        basic.get('\u8bfe\u9898\u627f\u62c5\u5355\u4f4d'),
        basic.get('\u8bfe\u9898\u7275\u5934\u627f\u62c5\u5355\u4f4d'),
        fields.get('\u8bfe\u9898\u7ec4\u7ec7\u5355\u4f4d')
    )
    ownership = _map_ownership(managing_unit)

    lead_name = _pick_first(
        basic.get('\u8bfe\u9898\u8d1f\u8d23\u4eba'),
        fields.get('\u59d3\u540d'),
        fields.get('\u8bfe\u9898\u8d1f\u8d23\u4eba')
    )
    if _is_placeholder_name(lead_name):
        lead_name = ''

    contact_person = _pick_first(
        basic.get('\u8bfe\u9898\u8054\u7cfb\u4eba'),
        fields.get('\u8bfe\u9898\u8054\u7cfb\u4eba')
    )

    start_date_text = _pick_first(fields.get('\u8d77\u59cb\u65f6\u95f4'))
    end_date_text = _pick_first(fields.get('\u7ec8\u6b62\u65f6\u95f4'))
    start_date = None
    end_date = None

    if start_date_text:
        dates = _find_dates(start_date_text)
        start_date = dates[0] if dates else None
    if end_date_text:
        dates = _find_dates(end_date_text)
        end_date = dates[0] if dates else None
    if not start_date and not end_date:
        date_range_text = _pick_first(basic.get('\u8bfe\u9898\u8d77\u6b62\u5e74\u9650'))
        start_date, end_date = _extract_date_range(date_range_text)

    budget_row = _extract_budget_totals(budget_rows)
    total_budget = _coerce_number(budget_row.get('\u5408\u8ba1')) if budget_row else None
    external_funding = _coerce_number(budget_row.get('\u4e13\u9879\u7ecf\u8d39')) if budget_row else None
    institute_funding = _coerce_number(budget_row.get('\u9662\u4e13\u9879\u7ecf\u8d39')) if budget_row else None
    unit_funding = _coerce_number(budget_row.get('\u81ea\u7b79\u7ecf\u8d39')) if budget_row else None
    if unit_funding is None and budget_row:
        unit_funding = _coerce_number(budget_row.get('\u6240\u5c5e\u5355\u4f4d\u81ea\u7b79\u8d44\u91d1'))

    funding_text = _pick_first(basic.get('\u7acb\u9879\u7ecf\u8d39'), fields.get('\u7ecf\u8d39\u9884\u7b97'))
    funding_values = _parse_funding_text(funding_text)
    if total_budget is None:
        total_budget = funding_values.get('total_budget')
    if institute_funding is None:
        institute_funding = funding_values.get('institute_funding')
    if external_funding is None:
        external_funding = funding_values.get('external_funding')
    if unit_funding is None:
        unit_funding = funding_values.get('unit_funding')
    if total_budget is None:
        parts = [v for v in (external_funding, institute_funding, unit_funding) if v is not None]
        if parts:
            total_budget = sum(parts)

    research_content = _normalize_text(fields.get('\u4e3b\u8981\u7814\u7a76\u5185\u5bb9'))

    mapped = {}
    if project_id:
        mapped['project_id'] = project_id
    if name:
        mapped['name'] = name
    if managing_unit:
        mapped['managing_unit'] = managing_unit
    if ownership:
        mapped['ownership'] = ownership
    if lead_name:
        mapped['project_lead'] = lead_name
    if contact_person:
        mapped['contact_person'] = contact_person
    if start_date:
        mapped['start_date'] = start_date.isoformat()
        mapped['start_year'] = str(start_date.year)
    if end_date:
        mapped['planned_end_date'] = end_date.isoformat()
    if total_budget is not None:
        mapped['total_budget'] = _format_number(total_budget)
    if external_funding is not None:
        mapped['external_funding'] = _format_number(external_funding)
    if institute_funding is not None:
        mapped['institute_funding'] = _format_number(institute_funding)
    if unit_funding is not None:
        mapped['unit_funding'] = _format_number(unit_funding)
    if research_content:
        mapped['research_content_manual'] = research_content
    return mapped

def _convert_doc_to_docx(doc_path):
    temp_dir = tempfile.mkdtemp()
    soffice_candidates = [
        shutil.which('soffice'),
        shutil.which('soffice.exe'),
        r'C:\Program Files\LibreOffice\program\soffice.exe',
        r'C:\Program Files (x86)\LibreOffice\program\soffice.exe',
    ]
    soffice_path = next((p for p in soffice_candidates if p and os.path.exists(p)), None)
    if not soffice_path:
        raise RuntimeError('未检测到LibreOffice，请安装后再解析.doc文件。')
    result = subprocess.run(
        [soffice_path, '--headless', '--convert-to', 'docx', '--outdir', temp_dir, doc_path],
        capture_output=True,
        text=True,
        encoding='utf-8',
        errors='ignore'
    )
    dest_path = os.path.join(temp_dir, f"{Path(doc_path).stem}.docx")
    if result.returncode != 0 or not os.path.exists(dest_path):
        candidates = list(Path(temp_dir).glob('*.docx'))
        if candidates:
            dest_path = str(candidates[0])
        else:
            error = (result.stderr or '').strip() or (result.stdout or '').strip() or 'DOC转DOCX失败'
            raise RuntimeError(error)
    return dest_path, temp_dir


@require_POST
def extract_task_docx_view(request):
    uploaded_file = request.FILES.get('document')
    if not uploaded_file:
        return JsonResponse({'success': False, 'message': '\u8bf7\u9009\u62e9\u8981\u89e3\u6790\u7684Word\u6587\u6863\u3002'})

    ext = os.path.splitext(uploaded_file.name)[1].lower()
    if ext not in ['.docx', '.doc']:
        return JsonResponse({'success': False, 'message': '\u4ec5\u652f\u6301.doc/.docx\u683c\u5f0f\u7684Word\u6587\u6863\u3002'})

    temp_path = None
    converted_path = None
    temp_dir = None
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=ext) as tmp:
            for chunk in uploaded_file.chunks():
                tmp.write(chunk)
            temp_path = tmp.name

        parse_path = temp_path
        if ext == '.doc':
            converted_path, temp_dir = _convert_doc_to_docx(temp_path)
            parse_path = converted_path

        extracted = extract_task_docx(parse_path)
        mapped = _map_docx_to_project_fields(extracted)
        return JsonResponse({'success': True, 'data': mapped})
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'\u89e3\u6790\u5931\u8d25: {e}'})
    finally:
        for path_to_remove in (temp_path, converted_path):
            if path_to_remove and os.path.exists(path_to_remove):
                try:
                    os.remove(path_to_remove)
                except OSError:
                    pass
        if temp_dir and os.path.isdir(temp_dir):
            try:
                shutil.rmtree(temp_dir)
            except OSError:
                pass

def project_list_view(request):
    queryset = _apply_project_filters(Project.objects.all(), request)
    distinct_years = Project.objects.values_list('start_year', flat=True).distinct().order_by('-start_year')
    distinct_statuses = Project.objects.values_list('status', flat=True).distinct().order_by('status')
    distinct_levels = Project.objects.values_list('level', flat=True).distinct().order_by('level')
    distinct_ownerships = Project.objects.values_list('ownership', flat=True).distinct().order_by('ownership')
    distinct_types = Project.objects.values_list('project_type', flat=True).distinct().order_by('project_type')
    distinct_roles = Project.objects.values_list('role', flat=True).distinct().order_by('role')
    distinct_units = Project.objects.exclude(managing_unit__isnull=True).exclude(managing_unit__exact='').values_list('managing_unit', flat=True).distinct().order_by('managing_unit')
    distinct_leads = Project.objects.exclude(project_lead__isnull=True).exclude(project_lead__exact='').values_list('project_lead', flat=True).distinct().order_by('project_lead')

    selected_years = _extract_list_param(request, 'year')
    selected_statuses = _extract_list_param(request, 'status')
    selected_levels = _extract_list_param(request, 'level')
    selected_ownerships = _extract_list_param(request, 'ownership')
    selected_types = _extract_list_param(request, 'project_type')
    selected_roles = _extract_list_param(request, 'role')
    selected_units = _extract_list_param(request, 'managing_unit')
    selected_leads = _extract_list_param(request, 'project_lead')
    min_budget = request.GET.get('min_budget', '').strip()
    max_budget = request.GET.get('max_budget', '').strip()
    start_date_from = request.GET.get('start_date_from', '').strip()
    start_date_to = request.GET.get('start_date_to', '').strip()
    end_date_from = request.GET.get('end_date_from', '').strip()
    end_date_to = request.GET.get('end_date_to', '').strip()
    query = request.GET.get('q', '').strip()
    filters_applied = any([
        query, selected_years, selected_statuses, selected_levels, selected_ownerships,
        selected_types, selected_roles, selected_units, selected_leads,
        min_budget, max_budget, start_date_from, start_date_to, end_date_from, end_date_to
    ])

    # Summary stats
    from django.db.models import Sum
    total_projects = queryset.count()
    completed_projects = queryset.filter(status__in=['未立项', '结题', '终止']).count()
    ongoing_projects = total_projects - completed_projects
    total_budget = queryset.aggregate(Sum('total_budget'))['total_budget__sum'] or 0

    context = {
        'projects': queryset,
        'distinct_years': distinct_years,
        'distinct_statuses': distinct_statuses,
        'distinct_levels': distinct_levels,
        'distinct_ownerships': distinct_ownerships,
        'distinct_types': distinct_types,
        'distinct_roles': distinct_roles,
        'distinct_units': distinct_units,
        'distinct_leads': distinct_leads,
        'total_projects': total_projects,
        'ongoing_projects': ongoing_projects,
        'completed_projects': completed_projects,
        'total_budget': total_budget,
        'selected_years': selected_years,
        'selected_statuses': selected_statuses,
        'selected_levels': selected_levels,
        'selected_ownerships': selected_ownerships,
        'selected_types': selected_types,
        'selected_roles': selected_roles,
        'selected_units': selected_units,
        'selected_leads': selected_leads,
        'min_budget': min_budget,
        'max_budget': max_budget,
        'start_date_from': start_date_from,
        'start_date_to': start_date_to,
        'end_date_from': end_date_from,
        'end_date_to': end_date_to,
        'query': query,
        'filters_applied': filters_applied,
    }
    return render(request, 'core/project_list.html', context)


def export_project_list_view(request):
    queryset = _apply_project_filters(Project.objects.all(), request)

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = '课题清单'

    headers = [
        '课题编号', '课题名称', '课题归属', '归口单位', '课题级别', '课题类型', '参与角色',
        '开始年份', '课题状态', '课题联系人', '课题负责人', '开始日期', '计划结束日期',
        '延期时间', '实际结题时间', '总预算(万元)', '外部专项经费(万元)', '院自筹经费(万元)',
        '所属单位自筹经费(万元)', '主要研究内容', '备注',
    ]
    worksheet.append(headers)

    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for project in queryset:
        worksheet.append([
            project.project_id,
            project.name,
            project.ownership,
            project.managing_unit,
            project.level,
            project.project_type,
            project.role,
            project.start_year,
            project.status,
            project.contact_person,
            project.project_lead,
            project.start_date.isoformat() if project.start_date else '',
            project.planned_end_date.isoformat() if project.planned_end_date else '',
            project.extension_date.isoformat() if project.extension_date else '',
            project.actual_completion_date.isoformat() if project.actual_completion_date else '',
            float(project.total_budget) if project.total_budget is not None else '',
            float(project.external_funding) if project.external_funding is not None else '',
            float(project.institute_funding) if project.institute_funding is not None else '',
            float(project.unit_funding) if project.unit_funding is not None else '',
            project.research_content,
            project.remarks,
        ])

    worksheet.freeze_panes = 'A2'
    worksheet.auto_filter.ref = worksheet.dimensions

    for column_cells in worksheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = max(len(str(cell.value or '')) for cell in column_cells)
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 40)
        if column_letter in {'T', 'U'}:
            for cell in column_cells[1:]:
                cell.alignment = Alignment(vertical='top', wrap_text=True)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    timestamp = timezone.localtime().strftime('%Y%m%d_%H%M%S')
    filename = f'课题清单_{timestamp}.xlsx'

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f"attachment; filename*=UTF-8''{quote(filename)}"
    return response

def progress_monitor_view(request):
    queryset = Project.objects.all()
    distinct_years = Project.objects.values_list('start_year', flat=True).distinct().order_by('-start_year')
    distinct_statuses = Project.objects.values_list('status', flat=True).distinct().order_by('status')
    progress_status_options = [
        {'value': 'ontrack', 'label': '\u672a\u5230\u4e2d\u671f'},
        {'value': 'midterm', 'label': '\u5df2\u5230\u4e2d\u671f'},
        {'value': 'overdue', 'label': '\u8d85\u671f'},
        {'value': 'pending', 'label': '\u672a\u5f00\u59cb'},
        {'value': 'unknown', 'label': '\u7f3a\u5c11\u65e5\u671f'},
        {'value': 'complete', 'label': '\u5df2\u5b8c\u6210'},
    ]
    valid_progress_status_values = {item['value'] for item in progress_status_options}

    query = request.GET.get('q')
    if query:
        queryset = queryset.filter(
            Q(name__icontains=query) |
            Q(project_id__icontains=query) |
            Q(project_lead__icontains=query) |
            Q(contact_person__icontains=query) |
            Q(ownership__icontains=query) |
            Q(managing_unit__icontains=query) |
            Q(level__icontains=query) |
            Q(project_type__icontains=query) |
            Q(research_content__icontains=query) |
            Q(remarks__icontains=query)
        )

    year = request.GET.get('year')
    if year:
        queryset = queryset.filter(start_year=year)

    status = request.GET.get('status')
    if status:
        queryset = queryset.filter(status=status)

    progress_status = (request.GET.get('progress_status') or '').strip()
    if progress_status not in valid_progress_status_values:
        progress_status = ''

    today = timezone.localdate()
    monitor_rows = []
    counts = {
        'total': 0,
        'ontrack': 0,
        'midterm': 0,
        'overdue': 0,
        'pending': 0,
        'unknown': 0,
        'complete': 0,
    }

    for project in queryset.order_by('-start_year', 'project_id'):
        node = _compute_progress_node(project, today)
        if progress_status and node['status_key'] != progress_status:
            continue
        monitor_rows.append({
            'project': project,
            **node,
        })
        counts['total'] += 1
        counts[node['status_key']] += 1

    context = {
        'monitor_rows': monitor_rows,
        'distinct_years': distinct_years,
        'distinct_statuses': distinct_statuses,
        'counts': counts,
        'today': today,
        'query': query or '',
        'selected_year': year or '',
        'selected_status': status or '',
        'progress_status_options': progress_status_options,
        'selected_progress_status': progress_status,
    }
    return render(request, 'core/progress_monitor.html', context)


def expense_monitor_view(request):
    sheet_name = 'Sheet2'
    file_path = Path(settings.BASE_DIR) / 'zichouktfeiyong' / '19-26.XLSX'
    threshold = _parse_threshold(request.GET.get('threshold'), default=0.85)
    expense_unit_divisor = 10000.0
    expense_unit_label = '万元'
    today = timezone.localdate()

    file_exists = file_path.exists()
    file_mtime = None
    if file_exists:
        try:
            file_mtime = datetime.fromtimestamp(file_path.stat().st_mtime, tz=timezone.get_current_timezone())
        except Exception:
            file_mtime = None

    company_groups = []
    unmatched_rows = []
    negative_rows = []
    over_budget_alerts = []
    total_rows = 0
    total_expense_sum = 0.0

    latest_import = ExpenseImport.objects.order_by('-created_at').first()
    comparison_base = latest_import
    imported = False
    import_message = None

    project_options = list(Project.objects.order_by('project_id').values_list('project_id', 'name'))
    mapping_entries = ExpenseMapping.objects.select_related('project').all()
    mapping_map = {m.normalized_text: m.project for m in mapping_entries if m.normalized_text}

    if not file_exists:
        messages.error(request, '找不到支出明细表，请确认文件路径存在。')
    else:
        try:
            import pandas as pd
        except Exception:
            messages.error(request, '无法读取Excel文件，请确认服务端已安装pandas。')
            pd = None

        if file_exists and pd is not None:
            try:
                df = pd.read_excel(file_path, sheet_name=sheet_name)
            except Exception as exc:
                messages.error(request, f'读取Excel失败: {exc}')
                df = None

            if df is not None:
                df.columns = [str(col).strip() for col in df.columns]
                required_cols = {'公司名称', '科研课题文本描述', '期末余额'}
                if not required_cols.issubset(df.columns):
                    missing = required_cols - set(df.columns)
                    messages.error(request, f'缺少必要列: {", ".join(sorted(missing))}')
                else:
                    df['期末余额'] = pd.to_numeric(df['期末余额'], errors='coerce').fillna(0)
                    records = df.to_dict('records')

                    projects = list(Project.objects.all())
                    candidates = []
                    for project in projects:
                        clean_name = _clean_match_text(project.name)
                        if clean_name:
                            candidates.append({'project': project, 'clean': clean_name})

                    aggregates = {}

                    for row in records:
                        company = str(row.get('公司名称') or '').strip()
                        description = str(row.get('科研课题文本描述') or '').strip()
                        amount_value = row.get('期末余额', 0)
                        try:
                            amount_raw = float(amount_value or 0)
                        except (TypeError, ValueError):
                            amount_raw = 0.0
                        amount = amount_raw / expense_unit_divisor if expense_unit_divisor else amount_raw

                        total_rows += 1
                        total_expense_sum += amount

                        if amount < 0:
                            negative_rows.append({
                                'company': company or '-'
                                , 'description': description or '-'
                                , 'amount': amount
                            })

                        if not description:
                            unmatched_rows.append({
                                'company': company or '-'
                                , 'description': '-'
                                , 'amount': amount
                                , 'best_name': '-'
                                , 'best_score': 0
                            })
                            continue

                        clean_desc = _clean_match_text(description)
                        if len(clean_desc) > 512:
                            clean_desc = clean_desc[:512]
                        best_score = 0.0
                        best_project = None
                        match_method = 'auto'
                        mapped_project = mapping_map.get(clean_desc)
                        if mapped_project:
                            best_project = mapped_project
                            best_score = 1.0
                            match_method = 'manual'
                        else:
                            for candidate in candidates:
                                score = _similarity_score(clean_desc, candidate['clean'])
                                if score > best_score:
                                    best_score = score
                                    best_project = candidate['project']
                                    if best_score == 1.0:
                                        break

                        if best_project and best_score >= threshold:
                            group_name = best_project.managing_unit or '未设置归口单位'
                            key = (group_name, best_project.project_id)
                            entry = aggregates.get(key)
                            if not entry:
                                entry = {
                                    'group_name': group_name,
                                    'project': best_project,
                                    'project_id': best_project.project_id,
                                    'project_name': best_project.name,
                                    'total': 0.0,
                                    'max_score': best_score,
                                    'sample_desc': description,
                                    'row_count': 0,
                                    'completion_date': _get_completion_date(best_project),
                                    'budget_value': float(best_project.total_budget or 0) if best_project.total_budget is not None else None,
                                    'over_budget': False,
                                    'over_amount': 0.0,
                                    'match_method': match_method,
                                }
                                aggregates[key] = entry
                            entry['total'] += amount
                            entry['row_count'] += 1
                            if best_score >= entry['max_score']:
                                entry['max_score'] = best_score
                                entry['sample_desc'] = description
                            if match_method == 'manual':
                                entry['match_method'] = 'manual'
                        else:
                            unmatched_rows.append({
                                'company': company or '-'
                                , 'description': description
                                , 'amount': amount
                                , 'best_name': best_project.name if best_project else '-'
                                , 'best_score': best_score
                            })

                    company_map = {}
                    for entry in aggregates.values():
                        budget_value = entry.get('budget_value')
                        if budget_value and entry['total'] > budget_value + 0.01:
                            entry['over_budget'] = True
                            entry['over_amount'] = entry['total'] - budget_value
                            over_budget_alerts.append({
                                'group_name': entry['group_name'],
                                'project_id': entry['project_id'],
                                'project_name': entry['project_name'],
                                'budget_value': budget_value,
                                'total': entry['total'],
                                'over_amount': entry['over_amount'],
                            })

                        company_map.setdefault(entry['group_name'], []).append(entry)

                    for group_name, items in sorted(company_map.items(), key=lambda x: x[0]):
                        items.sort(key=lambda x: x['total'], reverse=True)
                        company_groups.append({
                            'company': group_name,
                            'rows': items,
                            'total': sum(i['total'] for i in items),
                        })

                    should_save = False
                    if file_mtime and (latest_import is None or (latest_import.file_mtime and file_mtime > latest_import.file_mtime) or (latest_import.file_mtime is None)):
                        should_save = True

                    if should_save:
                        comparison_base = latest_import
                        with transaction.atomic():
                            new_import = ExpenseImport.objects.create(
                                source_file=str(file_path),
                                sheet_name=sheet_name,
                                file_mtime=file_mtime,
                                threshold=threshold,
                            )
                            snapshots = []
                            for entry in aggregates.values():
                                try:
                                    amount_decimal = Decimal(str(round(entry['total'], 2)))
                                except Exception:
                                    amount_decimal = Decimal('0')
                                snapshots.append(ExpenseSnapshot(
                                    import_log=new_import,
                                    project=entry['project'],
                                    project_name=entry['project_name'],
                                    company_name=entry['group_name'],
                                    matched_description=entry['sample_desc'] or '',
                                    match_score=entry['max_score'],
                                    total_expense=amount_decimal,
                                ))
                            if snapshots:
                                ExpenseSnapshot.objects.bulk_create(snapshots)

                        imported = True
                        latest_import = new_import
                        import_message = '已检测到数据更新，已自动生成本次支出快照。'

                    if import_message:
                        messages.info(request, import_message)

    growth_alerts = []
    comparison_time = None
    if comparison_base:
        comparison_time = comparison_base.created_at
        prev_map = {}
        for snap in comparison_base.snapshots.select_related('project'):
            if snap.project_id:
                prev_map[snap.project_id] = float(snap.total_expense or 0)

        for group in company_groups:
            for item in group['rows']:
                completion_date = item.get('completion_date')
                if completion_date and completion_date < today:
                    prev_total = prev_map.get(item['project_id'])
                    if prev_total is not None and item['total'] > prev_total + 0.01:
                        growth_alerts.append({
                            'group_name': group['company'],
                            'project_id': item['project_id'],
                            'project_name': item['project_name'],
                            'completion_date': completion_date,
                            'previous_total': prev_total,
                            'current_total': item['total'],
                            'increase': item['total'] - prev_total,
                        })

    unmatched_display = unmatched_rows[:200]
    negative_display = negative_rows[:200]

    context = {
        'company_groups': company_groups,
        'unmatched_rows': unmatched_display,
        'unmatched_total': len(unmatched_rows),
        'negative_rows': negative_display,
        'negative_total': len(negative_rows),
        'total_rows': total_rows,
        'total_expense_sum': total_expense_sum,
        'file_path': str(file_path),
        'sheet_name': sheet_name,
        'file_mtime': file_mtime,
        'threshold': threshold,
        'latest_import': latest_import,
        'comparison_time': comparison_time,
        'growth_alerts': growth_alerts,
        'imported': imported,
        'expense_unit_label': expense_unit_label,
        'expense_unit_divisor': expense_unit_divisor,
        'over_budget_alerts': over_budget_alerts,
        'over_budget_total': len(over_budget_alerts),
        'project_options': project_options,
    }
    return render(request, 'core/expense_monitor.html', context)


@require_POST
def expense_import_view(request):
    file_path = Path(settings.BASE_DIR) / 'zichouktfeiyong' / '19-26.XLSX'
    upload = request.FILES.get('expense_file')
    if not upload:
        messages.error(request, '请选择要上传的Excel文件。')
        return redirect('expense_monitor')

    ext = Path(upload.name).suffix.lower()
    if ext not in {'.xlsx', '.xls'}:
        messages.error(request, '仅支持Excel文件(.xlsx/.xls)。')
        return redirect('expense_monitor')

    try:
        file_path.parent.mkdir(parents=True, exist_ok=True)
        with open(file_path, 'wb') as handle:
            for chunk in upload.chunks():
                handle.write(chunk)
        messages.success(request, '支出明细表已上传，系统将自动刷新匹配结果。')
    except Exception as exc:
        messages.error(request, f'上传失败: {exc}')
    return redirect('expense_monitor')


@require_POST
def expense_mapping_view(request):
    description = request.POST.get('description', '').strip()
    project_id = request.POST.get('project_id', '').strip()
    if not description or not project_id:
        messages.error(request, '请填写描述并选择匹配课题。')
        return redirect('expense_monitor')

    project = get_object_or_404(Project, project_id=project_id)
    normalized = _clean_match_text(description)
    if len(normalized) > 512:
        normalized = normalized[:512]
    if not normalized:
        messages.error(request, '描述无法转换成匹配文本，请检查输入。')
        return redirect('expense_monitor')

    ExpenseMapping.objects.update_or_create(
        normalized_text=normalized,
        defaults={
            'description_text': description,
            'project': project,
        }
    )
    messages.success(request, f'已映射课题：{project.name}')
    redirect_url = reverse('expense_monitor')
    threshold = request.POST.get('threshold', '').strip()
    if threshold:
        redirect_url = f"{redirect_url}?threshold={threshold}"
    return redirect(redirect_url)


def delete_project_view(request, project_id):
    project = get_object_or_404(Project, project_id=project_id)
    if request.method == 'POST':
        if project.directory_path and os.path.exists(project.directory_path):
            try:
                shutil.rmtree(project.directory_path)
            except OSError as e:
                messages.error(request, f"删除文件夹失败: {e}")
                return redirect('project_detail', project_id=project.project_id)
        project.delete()
        messages.success(request, f"项目 {project.name} 已被成功删除。")
        return redirect('project_list')
    return render(request, 'core/delete_project.html', {'project': project})

def project_detail_view(request, project_id):
    project = get_object_or_404(Project, project_id=project_id)
    
    if request.method == 'POST' and 'document' not in request.FILES:
        form = ProjectForm(request.POST, instance=project)
        if form.is_valid():
            # Check if folder-naming fields have changed
            old_project_instance = Project.objects.get(pk=project_id)
            old_path = old_project_instance.directory_path
            
            updated_project = form.save(commit=False)
            
            if (old_project_instance.status != updated_project.status or 
                old_project_instance.start_year != updated_project.start_year or 
                old_project_instance.name != updated_project.name):
                rename_project_folder(request, updated_project, old_path)
            
            updated_project.save()
            form.save_m2m() # Save many-to-many fields if any

            return redirect('project_detail', project_id=updated_project.project_id)
    else:
        form = ProjectForm(instance=project)

    # 确保项目目录存在并获取文件树
    file_tree = []
    try:
        # 检查项目目录路径是否需要更新（确保与状态、年份、名称同步）
        expected_folder_name = _get_project_folder_name(project)
        expected_path = os.path.join(str(settings.PROJECTS_ROOT), expected_folder_name)
        
        # 如果目录路径与期望不符，需要更新
        if project.directory_path != expected_path:
            print(f"项目目录路径需要更新: {project.project_id}")  # 调试信息
            
            # 查找是否有旧目录需要重命名
            projects_root = str(settings.PROJECTS_ROOT)
            if os.path.exists(projects_root):
                for item in os.listdir(projects_root):
                    item_path = os.path.join(projects_root, item)
                    if (os.path.isdir(item_path) and 
                        item != expected_folder_name and 
                        project.project_id in item):
                        # 重命名旧目录
                        if not os.path.exists(expected_path):
                            os.rename(item_path, expected_path)
                            messages.info(request, f"项目目录已重命名以匹配当前状态")
                        break
            
            # 更新数据库中的路径
            project.directory_path = expected_path
            project.save(update_fields=['directory_path'])
        
        # 如果目录路径不存在或目录不存在，创建目录结构
        if not project.directory_path or not os.path.exists(project.directory_path):
            print(f"项目目录不存在，创建目录结构: {project.project_id}")  # 调试信息
            messages.info(request, f"正在为项目 {project.name} 创建标准目录结构...")
            create_project_directory_structure(project)
            # 重新获取项目实例以确保directory_path是最新的
            project.refresh_from_db()
        
        # 获取文件树
        if project.directory_path and os.path.exists(project.directory_path):
            # 确保子目录结构最新（包含目录重命名）
            create_project_directory_structure(project)
            print(f"开始获取文件树，目录路径: {project.directory_path}")  # 调试信息
            file_tree = get_directory_tree(project.directory_path, project.directory_path)
            print(f"文件树节点数: {len(file_tree)}")  # 调试信息
            
            if not file_tree:
                print(f"文件树为空，目录可能没有内容")  # 调试信息
                messages.info(request, "项目目录结构已创建，但暂无文件。您可以通过右侧文件操作面板上传文件。")
            
        else:
            print(f"无法获取文件树，目录路径: {project.directory_path}")  # 调试信息
            messages.warning(request, f"无法访问项目目录：{project.directory_path}")
            
    except Exception as e:
        print(f"处理项目目录时出错: {e}")  # 调试信息
        import traceback
        print(f"详细错误信息: {traceback.format_exc()}")  # 详细错误信息
        messages.error(request, f"处理项目目录时出错: {e}")

    # Calculate relative path for file uploads
    relative_directory_path = ""
    
    # Get analysis results
    research_analysis = project.analyses.filter(analysis_type='research_content').first()
    metrics_analysis = project.analyses.filter(analysis_type='output_metrics').first()
    
    # Get metrics items grouped by category
    metrics_items_by_category = {}
    if metrics_analysis:
        metrics_items = metrics_analysis.metrics_items.all()
        for item in metrics_items:
            category_display = item.get_category_display()
            if category_display not in metrics_items_by_category:
                metrics_items_by_category[category_display] = []
            metrics_items_by_category[category_display].append(item)
    
    # Get API configuration status
    api_configs = APIConfig.objects.filter(is_active=True)
    api_status = {
        'has_config': api_configs.exists(),
        'deepseek_available': api_configs.filter(service_name='deepseek', test_success=True).exists(),
        'kimi_available': api_configs.filter(service_name='kimi', test_success=True).exists(),
        'total_configs': api_configs.count()
    }

    network_config = get_network_config()
    protocol_setup_reg_path = str(Path(settings.BASE_DIR) / 'setup_protocol_handler.reg')
    client_setup_reg_path = str(Path(settings.BASE_DIR) / 'client_setup_protocol.reg')

    context = {
        'project': project,
        'form': form,
        'file_tree': file_tree,
        'research_analysis': research_analysis,
        'metrics_analysis': metrics_analysis,
        'metrics_items_by_category': metrics_items_by_category,
        'api_status': api_status,
        'relative_directory_path': relative_directory_path,
        'enable_network_share': network_config.get('enable_network_share', False),
        'network_share_path': network_config.get('network_share_path', ''),
        'enable_web_file_trial': network_config.get('enable_web_file_trial', True),
        'protocol_setup_reg_path': protocol_setup_reg_path,
        'client_setup_reg_path': client_setup_reg_path,
    }
    return render(request, 'core/project_detail.html', context)

def file_manager_trial_view(request, project_id):
    """Web 文件管理试用页，不影响现有模式。"""
    project = get_object_or_404(Project, project_id=project_id)
    network_config = get_network_config()

    if not network_config.get('enable_web_file_trial', True):
        messages.warning(request, 'Web 文件管理试用功能当前已关闭。')
        return redirect('project_detail', project_id=project_id)

    if not project.directory_path or not os.path.isdir(project.directory_path):
        create_project_directory_structure(project)
        project.refresh_from_db()

    if project.directory_path and os.path.exists(project.directory_path):
        create_project_directory_structure(project)

    return render(request, 'core/file_manager_trial.html', {
        'project': project,
    })

def get_file_tree_view(request, project_id):
    """API端点：返回项目文件树的JSON数据"""
    project = get_object_or_404(Project, project_id=project_id)
    
    # 确保项目目录存在
    if not project.directory_path or not os.path.isdir(project.directory_path):
        create_project_directory_structure(project)
    
    # 获取文件树数据
    file_tree = []
    if project.directory_path and os.path.exists(project.directory_path):
        # 确保子目录结构最新（包含目录重命名）
        create_project_directory_structure(project)
        file_tree = get_directory_tree(project.directory_path, project.directory_path)
    
    # 计算文件统计信息
    total_files = 0
    total_folders = 0
    
    def count_items(nodes):
        nonlocal total_files, total_folders
        for node in nodes:
            if node['type'] == 'folder':
                total_folders += 1
                if node.get('children'):
                    count_items(node['children'])
            else:
                total_files += 1
    
    count_items(file_tree)
    
    return JsonResponse({
        'success': True,
        'file_tree': file_tree,
        'stats': {
            'file_count': total_files,
            'folder_count': total_folders,
            'total_count': total_files + total_folders
        }
    })

def file_action_view(request, project_id, action):
    project = get_object_or_404(Project, project_id=project_id)
    if not project.directory_path or not os.path.isdir(project.directory_path):
        create_project_directory_structure(project)

    # Check if this is an AJAX request
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    project_root = project.directory_path

    if action == 'upload' and request.method == 'POST':
        print(f"[DEBUG] 文件上传请求 - 项目ID: {project_id}")
        print(f"[DEBUG] 是否AJAX请求: {is_ajax}")
        print(f"[DEBUG] POST数据: {dict(request.POST)}")
        print(f"[DEBUG] FILES数据: {list(request.FILES.keys())}")
        print(f"[DEBUG] request.FILES内容: {dict(request.FILES)}")
        
        uploaded_files = request.FILES.getlist('files')
        print(f"[DEBUG] getlist('files')结果: {uploaded_files}")
        target_path_rel = request.POST.get('target_path', '')
        
        print(f"[DEBUG] 上传文件数量: {len(uploaded_files)}")
        print(f"[DEBUG] 文件列表详情: {[f.name for f in uploaded_files]}")
        print(f"[DEBUG] 文件大小详情: {[f.size for f in uploaded_files]}")
        print(f"[DEBUG] 目标路径: {target_path_rel}")
        
        if not uploaded_files:
            print(f"[DEBUG] 没有找到上传文件")
            if is_ajax:
                return JsonResponse({'success': False, 'message': '请选择要上传的文件。'})
            messages.error(request, "请选择要上传的文件。")
            return redirect('project_detail', project_id=project.project_id)
        
        target_path_abs = _build_abs_path(project_root, target_path_rel)
        target_path_abs = os.path.normpath(target_path_abs)
        
        if not _is_within_root(target_path_abs, project_root):
            if is_ajax:
                return JsonResponse({'success': False, 'message': '无效的目标路径。'})
            messages.error(request, "无效的目标路径。")
            return redirect('project_detail', project_id=project.project_id)
        
        # 确保目标目录存在
        os.makedirs(target_path_abs, exist_ok=True)
        
        success_count = 0
        error_messages = []
        
        for uploaded_file in uploaded_files:
            try:
                print(f"[DEBUG] 处理文件: {uploaded_file.name}, 大小: {uploaded_file.size}")
                # 清理文件名，移除危险字符，但保留中文字符
                safe_filename = re.sub(r'[\\/*?"<>|:]', '_', uploaded_file.name)
                # 确保文件名不为空且不以点开头
                if not safe_filename or safe_filename.startswith('.'):
                    safe_filename = f"file_{int(time.time())}{os.path.splitext(uploaded_file.name)[1]}"
                file_path = os.path.join(target_path_abs, safe_filename)
                print(f"[DEBUG] 保存到: {file_path}")
                
                with open(file_path, 'wb+') as destination:
                    for chunk in uploaded_file.chunks():
                        destination.write(chunk)
                print(f"[DEBUG] 文件保存成功: {safe_filename}")
                success_count += 1
            except Exception as e:
                print(f"[DEBUG] 文件保存失败: {uploaded_file.name}, 错误: {str(e)}")
                error_messages.append(f"{uploaded_file.name}: {str(e)}")
        
        if success_count > 0:
            success_message = f"成功上传 {success_count} 个文件。"
            if error_messages:
                success_message += f" 失败: {'; '.join(error_messages)}"
            
            print(f"[DEBUG] 上传成功: {success_message}")
            if is_ajax:
                return JsonResponse({'success': True, 'message': success_message})
            messages.success(request, success_message)
        else:
            error_message = f"文件上传失败: {'; '.join(error_messages)}"
            print(f"[DEBUG] 上传失败: {error_message}")
            if is_ajax:
                return JsonResponse({'success': False, 'message': error_message})
            messages.error(request, error_message)

    elif action == 'delete':
        item_path_rel = request.POST.get('path') or request.GET.get('path')
        if item_path_rel:
            item_path_abs = _build_abs_path(project_root, item_path_rel)
            item_path_abs = os.path.normpath(item_path_abs)
            
            if not _is_within_root(item_path_abs, project_root):
                if is_ajax:
                    return JsonResponse({'success': False, 'message': '无效的文件路径。'})
                messages.error(request, "无效的文件路径。")
                return redirect('project_detail', project_id=project.project_id)
                
            if os.path.exists(item_path_abs):
                try:
                    if os.path.isdir(item_path_abs):
                        shutil.rmtree(item_path_abs)
                        if is_ajax:
                            return JsonResponse({'success': True, 'message': '文件夹删除成功！'})
                        messages.success(request, "文件夹删除成功！")
                    else:
                        os.remove(item_path_abs)
                        if is_ajax:
                            return JsonResponse({'success': True, 'message': '文件删除成功！'})
                        messages.success(request, "文件删除成功！")
                except OSError as e:
                    if is_ajax:
                        return JsonResponse({'success': False, 'message': f'删除失败: {e}'})
                    messages.error(request, f"删除失败: {e}")
            else:
                if is_ajax:
                    return JsonResponse({'success': False, 'message': '文件或文件夹不存在。'})
                messages.error(request, "文件或文件夹不存在。")

    elif action == 'rename' and request.method == 'POST':
        item_path_rel = request.POST.get('path', '')
        new_name = request.POST.get('new_name', '').strip()

        if not item_path_rel:
            if is_ajax:
                return JsonResponse({'success': False, 'message': '缺少目标路径。'})
            messages.error(request, "缺少目标路径。")
            return redirect('project_detail', project_id=project.project_id)

        if not new_name:
            if is_ajax:
                return JsonResponse({'success': False, 'message': '新名称不能为空。'})
            messages.error(request, "新名称不能为空。")
            return redirect('project_detail', project_id=project.project_id)

        safe_name = re.sub(r'[\\/*?"<>|:]', '_', new_name)
        if safe_name in ('.', '..'):
            if is_ajax:
                return JsonResponse({'success': False, 'message': '新名称不合法。'})
            messages.error(request, "新名称不合法。")
            return redirect('project_detail', project_id=project.project_id)

        item_path_abs = _build_abs_path(project_root, item_path_rel)
        item_path_abs = os.path.normpath(item_path_abs)
        if not _is_within_root(item_path_abs, project_root):
            if is_ajax:
                return JsonResponse({'success': False, 'message': '无效的目标路径。'})
            messages.error(request, "无效的目标路径。")
            return redirect('project_detail', project_id=project.project_id)

        if not os.path.exists(item_path_abs):
            if is_ajax:
                return JsonResponse({'success': False, 'message': '目标不存在。'})
            messages.error(request, "目标不存在。")
            return redirect('project_detail', project_id=project.project_id)

        parent_dir = os.path.dirname(item_path_abs)
        new_path_abs = os.path.normpath(os.path.join(parent_dir, safe_name))
        if not _is_within_root(new_path_abs, project_root):
            if is_ajax:
                return JsonResponse({'success': False, 'message': '重命名目标非法。'})
            messages.error(request, "重命名目标非法。")
            return redirect('project_detail', project_id=project.project_id)

        if os.path.exists(new_path_abs):
            if is_ajax:
                return JsonResponse({'success': False, 'message': '同名文件或文件夹已存在。'})
            messages.error(request, "同名文件或文件夹已存在。")
            return redirect('project_detail', project_id=project.project_id)

        try:
            os.rename(item_path_abs, new_path_abs)
            if is_ajax:
                return JsonResponse({'success': True, 'message': '重命名成功。'})
            messages.success(request, "重命名成功。")
        except OSError as e:
            if is_ajax:
                return JsonResponse({'success': False, 'message': f'重命名失败: {e}'})
            messages.error(request, f"重命名失败: {e}")
    
    elif action == 'download':
        item_path_rel = request.GET.get('path')
        if item_path_rel:
            item_path_abs = _build_abs_path(project_root, item_path_rel)
            item_path_abs = os.path.normpath(item_path_abs)
            
            if not _is_within_root(item_path_abs, project_root):
                messages.error(request, "无效的文件路径。")
                return redirect('project_detail', project_id=project.project_id)
                
            if os.path.exists(item_path_abs) and os.path.isfile(item_path_abs):
                try:
                    response = FileResponse(
                        open(item_path_abs, 'rb'),
                        as_attachment=True,
                        filename=force_str(os.path.basename(item_path_abs))
                    )
                    return response
                except Exception as e:
                    messages.error(request, f"文件下载失败: {e}")
            else:
                messages.error(request, "文件不存在。")
    
    elif action == 'preview':
        item_path_rel = request.GET.get('path')
        if item_path_rel:
            item_path_abs = _build_abs_path(project_root, item_path_rel)
            item_path_abs = os.path.normpath(item_path_abs)
            
            if not _is_within_root(item_path_abs, project_root):
                return HttpResponse("无效的文件路径", status=403)
                
            if os.path.exists(item_path_abs) and os.path.isfile(item_path_abs):
                try:
                    # Get file content type
                    content_type, _ = mimetypes.guess_type(item_path_abs)
                    
                    # For text files, try to read and display content
                    if content_type and content_type.startswith('text/'):
                        try:
                            with open(item_path_abs, 'r', encoding='utf-8') as f:
                                content = f.read()
                            return HttpResponse(content, content_type='text/plain; charset=utf-8')
                        except UnicodeDecodeError:
                            # Try with other encodings
                            try:
                                with open(item_path_abs, 'r', encoding='gbk') as f:
                                    content = f.read()
                                return HttpResponse(content, content_type='text/plain; charset=utf-8')
                            except UnicodeDecodeError:
                                return HttpResponse("无法解码文件内容", status=400)
                    
                    # For images, return the file directly
                    elif content_type and content_type.startswith('image/'):
                        return FileResponse(open(item_path_abs, 'rb'), content_type=content_type)
                    
                    # For PDFs, return the file directly
                    elif content_type == 'application/pdf':
                        return FileResponse(open(item_path_abs, 'rb'), content_type=content_type)
                    
                    else:
                        return HttpResponse(f"不支持预览此文件类型: {content_type or '未知类型'}", status=400)
                        
                except Exception as e:
                    return HttpResponse(f"文件预览失败: {e}", status=500)
            else:
                return HttpResponse("文件不存在", status=404)
    
    elif action == 'create_folder' and request.method == 'POST':
        folder_name = request.POST.get('folder_name', '').strip()
        parent_path_rel = request.POST.get('parent_path', '')
        
        if not folder_name:
            if is_ajax:
                return JsonResponse({'success': False, 'message': '文件夹名称不能为空。'})
            messages.error(request, "文件夹名称不能为空。")
            return redirect('project_detail', project_id=project.project_id)
        
        # Sanitize folder name
        folder_name = re.sub(r'[\\/*?"<>|]', '_', folder_name)
        
        parent_path_abs = _build_abs_path(project_root, parent_path_rel)
        parent_path_abs = os.path.normpath(parent_path_abs)
        
        if not _is_within_root(parent_path_abs, project_root):
            if is_ajax:
                return JsonResponse({'success': False, 'message': '无效的父目录路径。'})
            messages.error(request, "无效的父目录路径。")
            return redirect('project_detail', project_id=project.project_id)
        
        new_folder_path = os.path.join(parent_path_abs, folder_name)
        
        if os.path.exists(new_folder_path):
            if is_ajax:
                return JsonResponse({'success': False, 'message': f'文件夹 \'{folder_name}\' 已存在。'})
            messages.error(request, f"文件夹 '{folder_name}' 已存在。")
        else:
            try:
                os.makedirs(new_folder_path, exist_ok=True)
                if is_ajax:
                    return JsonResponse({'success': True, 'message': f'文件夹 \'{folder_name}\' 创建成功！'})
                messages.success(request, f"文件夹 '{folder_name}' 创建成功！")
            except Exception as e:
                if is_ajax:
                    return JsonResponse({'success': False, 'message': f'创建文件夹失败: {e}'})
                messages.error(request, f"创建文件夹失败: {e}")

    return redirect('project_detail', project_id=project.project_id)

def analyze_content_view(request, project_id, analysis_type):
    from .ai_analysis import ai_service
    
    project = get_object_or_404(Project, project_id=project_id)
    
    if request.method == 'POST':
        uploaded_file = request.FILES.get('document')
        
        if uploaded_file:
            # 使用AI分析服务处理文档
            result = ai_service.analyze_document(uploaded_file, analysis_type)
            
            if result['success']:
                # 保存或更新分析结果
                analysis, created = ProjectAnalysis.objects.update_or_create(
                    project=project,
                    analysis_type=analysis_type,
                    defaults={
                        'file_name': uploaded_file.name,
                        'file_size': uploaded_file.size,
                        'analysis_result': result['result'],
                        'confidence_score': result.get('confidence_score'),
                        'processing_time': result['processing_time'],
                    }
                )
                
                action = '创建' if created else '更新'
                api_info = f" (使用{result.get('api_used', 'AI')}服务)" if result.get('api_used') else ""
                messages.success(request, f'文档分析完成！{action}了{analysis.get_analysis_type_display()}结果。{api_info}')
            else:
                # 分析失败，显示错误信息
                messages.error(request, f'文档分析失败: {result["error"]}')
        else:
            messages.error(request, '请选择要分析的文件。')
    
    return redirect('project_detail', project_id=project.project_id)

def parse_metrics_analysis(analysis_text):
    """解析产出指标分析文本，提取指标项目"""
    metrics_items = []
    
    # 定义指标分类映射
    category_mapping = {
        '技术成果': 'technical',
        '学术成果': 'academic', 
        '标准制定': 'standard',
        '人才培养': 'talent',
        '经济效益': 'economic'
    }
    
    lines = analysis_text.split('\n')
    current_category = None
    
    for line in lines:
        line = line.strip()
        if not line or line.startswith('---') or line.startswith('###'):
            continue
            
        # 检查是否是分类标题
        for category_name, category_code in category_mapping.items():
            if category_name in line and ('指标' in line or '成果' in line or '效益' in line):
                current_category = category_code
                break
        
        # 解析指标项目
        if current_category and line.startswith('-'):
            # 移除开头的 '-' 和空格
            item_text = line.lstrip('- ').strip()
            if ':' in item_text or '：' in item_text:
                # 分割项目名称和目标值
                if '：' in item_text:
                    parts = item_text.split('：', 1)
                else:
                    parts = item_text.split(':', 1)
                
                if len(parts) == 2:
                    item_name = parts[0].strip()
                    target_value = parts[1].strip()
                    
                    # 提取括号中的详细信息作为备注
                    notes = ''
                    if '（' in target_value and '）' in target_value:
                        start = target_value.find('（')
                        end = target_value.find('）') + 1
                        notes = target_value[start:end]
                        target_value = target_value[:start].strip()
                    elif '(' in target_value and ')' in target_value:
                        start = target_value.find('(')
                        end = target_value.find(')') + 1
                        notes = target_value[start:end]
                        target_value = target_value[:start].strip()
                    
                    metrics_items.append({
                        'category': current_category,
                        'item_name': item_name,
                        'target_value': target_value,
                        'notes': notes
                    })
    
    return metrics_items


def edit_analysis_view(request, project_id, analysis_type):
    """编辑分析结果的视图"""
    project = get_object_or_404(Project, project_id=project_id)
    
    if request.method == 'POST':
        analysis_result = request.POST.get('analysis_result', '').strip()
        
        if analysis_result:
            # 创建或更新分析结果
            analysis, created = ProjectAnalysis.objects.update_or_create(
                project=project,
                analysis_type=analysis_type,
                defaults={
                    'file_name': '手动编辑',
                    'file_size': None,
                    'analysis_result': analysis_result,
                    'confidence_score': None,  # 手动编辑没有置信度
                    'processing_time': None,
                }
            )
            
            # 如果是产出指标分析，解析并创建指标项目
            if analysis_type == 'output_metrics':
                # 删除现有的指标项目
                analysis.metrics_items.all().delete()
                
                # 解析新的指标项目
                metrics_items = parse_metrics_analysis(analysis_result)
                
                # 创建新的指标项目记录
                for item_data in metrics_items:
                    MetricsItem.objects.create(
                        analysis=analysis,
                        category=item_data['category'],
                        item_name=item_data['item_name'],
                        target_value=item_data['target_value'],
                        notes=item_data['notes'],
                        status='pending'  # 默认状态为待完成
                    )
                
                messages.success(request, f'手动{"创建" if created else "更新"}了{analysis.get_analysis_type_display()}结果，解析出 {len(metrics_items)} 个指标项目。')
            else:
                action = '创建' if created else '更新'
                messages.success(request, f'手动{action}了{analysis.get_analysis_type_display()}结果。')
        else:
            messages.error(request, '分析结果内容不能为空。')
    
    return redirect('project_detail', project_id=project.project_id)


def update_metrics_item_view(request, project_id, item_id):
    """更新指标项目状态的视图"""
    if request.method == 'POST':
        item = get_object_or_404(MetricsItem, id=item_id, analysis__project__project_id=project_id)
        
        status = request.POST.get('status')
        current_value = request.POST.get('current_value', '')
        notes = request.POST.get('notes', '')
        
        if status in dict(MetricsItem.STATUS_CHOICES):
            item.status = status
            item.current_value = current_value
            item.notes = notes
            item.save()
            
            messages.success(request, f'已更新指标项目 "{item.item_name}" 的状态。')
        else:
            messages.error(request, '无效的状态值。')
    
    return redirect('project_detail', project_id=project_id)


def import_from_excel_view(request):
    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')
        if not excel_file:
            messages.error(request, "请选择要上传的Excel文件。")
            return redirect('project_list')

        try:
            workbook = openpyxl.load_workbook(excel_file)
            sheet = workbook.active
            header = [cell.value for cell in sheet[1]]
            print(f"Excel表头: {header}")  # 调试信息
            field_mapping = {
                '序号': None,  # 跳过序号列
                '课题编号': 'project_id', '课题名称': 'name', '课题归属': 'ownership',
                '归口单位': 'managing_unit', '课题级别': 'level', '课题类型': 'project_type',
                '参与角色': 'role', '开始年份': 'start_year', '课题状态': 'status',
                '课题联系人': 'contact_person', '课题负责人': 'project_lead', '开始日期': 'start_date',
                '计划结束日期': 'planned_end_date', '延期时间': 'extension_date', '实际结题时间': 'actual_completion_date',
                '总预算': 'total_budget', '外部专项经费': 'external_funding',
                '院自筹经费': 'institute_funding', '所属单位自筹经费': 'unit_funding',
                '主要研究内容': 'research_content', '备注': 'remarks',
            }

            processed_count = 0
            created_count = 0
            updated_count = 0
            skipped_count = 0
            auto_filled_count = 0
            for row in sheet.iter_rows(min_row=2, values_only=True):
                project_data = dict(zip(header, row))
                print(f"处理行数据: {project_data}")  # 调试信息
                model_data = {}
                for header_name, model_field in field_mapping.items():
                    if model_field is None:  # 跳过不需要的字段（如序号）
                        continue
                    if header_name in project_data and project_data[header_name] is not None:
                        value = project_data[header_name]
                        
                        # 处理日期字段
                        if model_field in ['start_date', 'planned_end_date', 'extension_date', 'actual_completion_date']:
                            if isinstance(value, str) and value.strip():
                                try:
                                    from datetime import datetime
                                    # 尝试多种日期格式
                                    for date_format in ['%Y-%m-%d', '%Y/%m/%d', '%Y.%m.%d', '%Y年%m月%d日']:
                                        try:
                                            model_data[model_field] = datetime.strptime(value.strip(), date_format).date()
                                            break
                                        except ValueError:
                                            continue
                                    else:
                                        # 如果所有格式都失败，跳过这个字段
                                        continue
                                except Exception:
                                    continue
                            elif hasattr(value, 'date'):  # Excel日期对象
                                try:
                                    model_data[model_field] = value.date()
                                except Exception:
                                    continue
                        
                        # 处理预算字段
                        elif model_field in ['total_budget', 'external_funding', 'institute_funding', 'unit_funding']:
                            if isinstance(value, (int, float)) and value > 0:
                                model_data[model_field] = value
                            elif isinstance(value, str) and value.strip():
                                try:
                                    # 移除可能的货币符号和空格
                                    clean_value = value.strip().replace('万元', '').replace('万', '').replace('元', '').replace(',', '').replace('，', '')
                                    if clean_value:
                                        model_data[model_field] = float(clean_value)
                                except (ValueError, TypeError):
                                    continue

                        # 处理开始年份
                        elif model_field == 'start_year':
                            if isinstance(value, (int, float)):
                                try:
                                    model_data[model_field] = int(value)
                                except (ValueError, TypeError):
                                    continue
                            elif isinstance(value, str) and value.strip():
                                match = re.search(r'\d{4}', value.strip())
                                if match:
                                    model_data[model_field] = int(match.group())
                                else:
                                    continue
                        
                        # 处理其他字段
                        else:
                            model_data[model_field] = value

                project_id = model_data.get('project_id')
                if not project_id:
                    print(f"跳过空项目ID的行")
                    skipped_count += 1
                    continue

                existing_project = Project.objects.filter(project_id=project_id).first()
                is_new_project = existing_project is None

                if is_new_project:
                    missing_fields = []
                    required_fields = {
                        'name': '课题名称',
                        'ownership': '课题归属',
                        'level': '课题级别',
                        'project_type': '课题类型',
                        'role': '参与角色',
                        'start_year': '开始年份',
                        'status': '课题状态',
                    }

                    for field, label in required_fields.items():
                        value = model_data.get(field)
                        is_missing = False
                        if field == 'start_year':
                            if value in (None, '') or (isinstance(value, (int, float)) and int(value) == 0):
                                is_missing = True
                        else:
                            if value is None or (isinstance(value, str) and not value.strip()):
                                is_missing = True

                        if is_missing:
                            missing_fields.append(label)
                            if field == 'name':
                                model_data[field] = f"待补全-{project_id}"
                            elif field == 'start_year':
                                model_data[field] = timezone.now().year
                            else:
                                model_data[field] = '待补全'

                    if missing_fields:
                        auto_note = f"【导入待补全】缺失字段：{'、'.join(missing_fields)}"
                        existing_remarks = model_data.get('remarks', '')
                        if existing_remarks and isinstance(existing_remarks, str):
                            model_data['remarks'] = f"{existing_remarks}\n{auto_note}"
                        else:
                            model_data['remarks'] = auto_note
                        auto_filled_count += 1

                    if 'directory_path' not in model_data:
                        model_data['directory_path'] = ''

                print(f"准备创建/更新项目: {project_id}, 数据: {model_data}")
                project, created = Project.objects.update_or_create(
                    project_id=project_id,
                    defaults=model_data
                )
                create_project_directory_structure(project)
                processed_count += 1
                if created:
                    created_count += 1
                else:
                    updated_count += 1
                print(f"{'创建' if created else '更新'}项目: {project_id}")

            print(f"导入完成，共处理 {processed_count} 个项目")
            messages.success(
                request,
                f"数据导入成功！共处理 {processed_count} 个项目（新建 {created_count}，更新 {updated_count}，跳过 {skipped_count}）。"
            )
            if auto_filled_count:
                messages.warning(
                    request,
                    f"其中 {auto_filled_count} 个新项目存在字段缺失，已用“待补全”自动填充，请在详情页补全。"
                )
        except Exception as e:
            import traceback
            error_msg = f"处理文件时出错: {e}\n{traceback.format_exc()}"
            print(error_msg)  # 输出到终端
            messages.error(request, f"处理文件时出错: {e}")

        return redirect('project_list')
    
    return redirect('project_list')

def statistics_view(request):
    from django.db.models import Sum, Avg
    from django.db.models import Case, When, IntegerField

    base_queryset = Project.objects.all()
    filtered_queryset = _apply_project_filters(base_queryset, request)

    status_distribution = list(filtered_queryset.values('status').annotate(count=Count('status')).order_by('-count'))
    level_distribution = list(filtered_queryset.values('level').annotate(count=Count('level')).order_by('-count'))
    yearly_distribution = list(filtered_queryset.values('start_year').annotate(count=Count('start_year')).order_by('start_year'))
    ownership_distribution = list(filtered_queryset.values('ownership').annotate(count=Count('ownership')).order_by('-count'))
    type_distribution = list(filtered_queryset.values('project_type').annotate(count=Count('project_type')).order_by('-count'))
    role_distribution = list(filtered_queryset.values('role').annotate(count=Count('role')).order_by('-count'))

    metrics_queryset = MetricsItem.objects.filter(analysis__project__in=filtered_queryset)
    metrics_category_distribution = list(metrics_queryset.values('category').annotate(count=Count('category')).order_by('-count'))
    metrics_status_distribution = list(metrics_queryset.values('status').annotate(count=Count('status')).order_by('-count'))

    status_count_dict = {item['status']: item['count'] for item in metrics_status_distribution}
    total_metrics = sum(status_count_dict.values())
    completed_metrics = status_count_dict.get('completed', 0)
    in_progress_metrics = status_count_dict.get('in_progress', 0)
    pending_metrics = status_count_dict.get('pending', 0)

    category_stats = metrics_queryset.values('category').annotate(
        total=Count('id'),
        completed=Count(Case(When(status='completed', then=1), output_field=IntegerField()))
    )
    category_stats_dict = {item['category']: item for item in category_stats}

    metrics_completion_by_category = []
    for category_code, category_name in MetricsItem.CATEGORY_CHOICES:
        stats = category_stats_dict.get(category_code, {'total': 0, 'completed': 0})
        category_total = stats['total']
        category_completed = stats['completed']
        completion_rate = (category_completed / category_total * 100) if category_total > 0 else 0

        metrics_completion_by_category.append({
            'category': category_name,
            'total': category_total,
            'completed': category_completed,
            'completion_rate': completion_rate
        })

    budget_stats = filtered_queryset.aggregate(
        total_budget_sum=Sum('total_budget'),
        external_funding_sum=Sum('external_funding'),
        institute_funding_sum=Sum('institute_funding'),
        unit_funding_sum=Sum('unit_funding'),
        avg_total_budget=Avg('total_budget')
    )

    yearly_budget = list(filtered_queryset.values('start_year').annotate(
        total_budget_sum=Sum('total_budget'),
        count=Count('project_id')
    ).order_by('start_year'))

    status_budget = list(filtered_queryset.values('status').annotate(
        total_budget_sum=Sum('total_budget'),
        count=Count('project_id')
    ).order_by('-total_budget_sum'))

    level_budget = list(filtered_queryset.values('level').annotate(
        total_budget_sum=Sum('total_budget'),
        count=Count('project_id')
    ).order_by('-total_budget_sum'))

    total_projects = filtered_queryset.count()
    completed_projects = filtered_queryset.filter(status__in=['未立项', '结题', '终止']).count()
    ongoing_projects = total_projects - completed_projects

    distinct_years = base_queryset.values_list('start_year', flat=True).distinct().order_by('-start_year')
    distinct_statuses = base_queryset.values_list('status', flat=True).distinct().order_by('status')
    distinct_levels = base_queryset.values_list('level', flat=True).distinct().order_by('level')
    distinct_ownerships = base_queryset.values_list('ownership', flat=True).distinct().order_by('ownership')
    distinct_types = base_queryset.values_list('project_type', flat=True).distinct().order_by('project_type')
    distinct_roles = base_queryset.values_list('role', flat=True).distinct().order_by('role')
    distinct_units = base_queryset.exclude(managing_unit__isnull=True).exclude(managing_unit__exact='').values_list('managing_unit', flat=True).distinct().order_by('managing_unit')
    distinct_leads = base_queryset.exclude(project_lead__isnull=True).exclude(project_lead__exact='').values_list('project_lead', flat=True).distinct().order_by('project_lead')

    selected_years = _extract_list_param(request, 'year')
    selected_statuses = _extract_list_param(request, 'status')
    selected_levels = _extract_list_param(request, 'level')
    selected_ownerships = _extract_list_param(request, 'ownership')
    selected_types = _extract_list_param(request, 'project_type')
    selected_roles = _extract_list_param(request, 'role')
    selected_units = _extract_list_param(request, 'managing_unit')
    selected_leads = _extract_list_param(request, 'project_lead')
    min_budget = request.GET.get('min_budget', '').strip()
    max_budget = request.GET.get('max_budget', '').strip()
    start_date_from = request.GET.get('start_date_from', '').strip()
    start_date_to = request.GET.get('start_date_to', '').strip()
    end_date_from = request.GET.get('end_date_from', '').strip()
    end_date_to = request.GET.get('end_date_to', '').strip()
    query = request.GET.get('q', '').strip()
    filters_applied = any([
        query, selected_years, selected_statuses, selected_levels, selected_ownerships,
        selected_types, selected_roles, selected_units, selected_leads,
        min_budget, max_budget, start_date_from, start_date_to, end_date_from, end_date_to
    ])

    context = {
        'status_data': json.dumps({
            'labels': [item['status'] for item in status_distribution],
            'data': [item['count'] for item in status_distribution],
        }),
        'level_data': json.dumps({
            'labels': [item['level'] for item in level_distribution],
            'data': [item['count'] for item in level_distribution],
        }),
        'yearly_data': json.dumps({
            'labels': [str(item['start_year']) for item in yearly_distribution],
            'data': [item['count'] for item in yearly_distribution],
        }),
        'ownership_data': json.dumps({
            'labels': [item['ownership'] for item in ownership_distribution],
            'data': [item['count'] for item in ownership_distribution],
        }),
        'type_data': json.dumps({
            'labels': [item['project_type'] for item in type_distribution],
            'data': [item['count'] for item in type_distribution],
        }),
        'role_data': json.dumps({
            'labels': [item['role'] for item in role_distribution],
            'data': [item['count'] for item in role_distribution],
        }),
        'yearly_budget_data': json.dumps({
            'labels': [str(item['start_year']) for item in yearly_budget],
            'data': [float(item['total_budget_sum'] or 0) for item in yearly_budget],
        }),
        'status_budget_data': json.dumps({
            'labels': [item['status'] for item in status_budget],
            'data': [float(item['total_budget_sum'] or 0) for item in status_budget],
        }),
        'level_budget_data': json.dumps({
            'labels': [item['level'] for item in level_budget],
            'data': [float(item['total_budget_sum'] or 0) for item in level_budget],
        }),
        'budget_stats': budget_stats,
        'total_projects': total_projects,
        'ongoing_projects': ongoing_projects,
        'completed_projects': completed_projects,
        'metrics_category_data': json.dumps({
            'labels': [dict(MetricsItem.CATEGORY_CHOICES).get(item['category'], item['category']) for item in metrics_category_distribution],
            'data': [item['count'] for item in metrics_category_distribution],
        }),
        'metrics_status_data': json.dumps({
            'labels': [dict(MetricsItem.STATUS_CHOICES).get(item['status'], item['status']) for item in metrics_status_distribution],
            'data': [item['count'] for item in metrics_status_distribution],
        }),
        'metrics_completion_data': json.dumps({
            'labels': [item['category'] for item in metrics_completion_by_category],
            'completion_rates': [item['completion_rate'] for item in metrics_completion_by_category],
            'totals': [item['total'] for item in metrics_completion_by_category],
            'completed': [item['completed'] for item in metrics_completion_by_category],
        }),
        'total_metrics': total_metrics,
        'completed_metrics': completed_metrics,
        'in_progress_metrics': in_progress_metrics,
        'pending_metrics': pending_metrics,
        'metrics_completion_by_category': metrics_completion_by_category,
        'distinct_years': distinct_years,
        'distinct_statuses': distinct_statuses,
        'distinct_levels': distinct_levels,
        'distinct_ownerships': distinct_ownerships,
        'distinct_types': distinct_types,
        'distinct_roles': distinct_roles,
        'distinct_units': distinct_units,
        'distinct_leads': distinct_leads,
        'selected_years': selected_years,
        'selected_statuses': selected_statuses,
        'selected_levels': selected_levels,
        'selected_ownerships': selected_ownerships,
        'selected_types': selected_types,
        'selected_roles': selected_roles,
        'selected_units': selected_units,
        'selected_leads': selected_leads,
        'min_budget': min_budget,
        'max_budget': max_budget,
        'start_date_from': start_date_from,
        'start_date_to': start_date_to,
        'end_date_from': end_date_from,
        'end_date_to': end_date_to,
        'query': query,
        'filters_applied': filters_applied,
    }
    return render(request, 'core/statistics.html', context)

def api_config_view(request):
    """API配置管理页面"""
    
    # 获取现有配置
    configs = APIConfig.objects.all().order_by('service_name')
    
    if request.method == 'POST':
        service_name = request.POST.get('service_name')
        api_key = request.POST.get('api_key', '').strip()
        action = request.POST.get('action')
        
        if action == 'save' and service_name and api_key:
            try:
                # 创建或更新配置
                config, created = APIConfig.objects.update_or_create(
                    service_name=service_name,
                    defaults={
                        'is_active': True,
                        'test_success': False,
                        'last_test_time': None
                    }
                )
                
                # 设置API密钥（自动加密）
                config.set_api_key(api_key)
                config.save()
                
                action_text = '创建' if created else '更新'
                messages.success(request, f'{config.get_service_name_display()} API配置{action_text}成功！')
                
            except Exception as e:
                messages.error(request, f'保存API配置失败: {e}')
        
        elif action == 'test' and service_name:
            try:
                config = APIConfig.objects.get(service_name=service_name)
                api_key = config.get_api_key()
                
                if not api_key:
                    messages.error(request, 'API密钥为空，无法测试')
                else:
                    # 测试API连接
                    test_success = test_api_connection(service_name, api_key)
                    
                    # 更新测试结果
                    config.test_success = test_success
                    config.last_test_time = timezone.now()
                    config.save()
                    
                    if test_success:
                        messages.success(request, f'{config.get_service_name_display()} API连接测试成功！')
                    else:
                        messages.error(request, f'{config.get_service_name_display()} API连接测试失败，请检查密钥是否正确')
                        
            except APIConfig.DoesNotExist:
                messages.error(request, '请先保存API配置再进行测试')
            except Exception as e:
                messages.error(request, f'API测试失败: {e}')
        
        elif action == 'toggle' and service_name:
            try:
                config = APIConfig.objects.get(service_name=service_name)
                config.is_active = not config.is_active
                config.save()
                
                status_text = '启用' if config.is_active else '禁用'
                messages.info(request, f'{config.get_service_name_display()} 服务已{status_text}')
                
            except APIConfig.DoesNotExist:
                messages.error(request, 'API配置不存在')
        
        elif action == 'delete' and service_name:
            try:
                config = APIConfig.objects.get(service_name=service_name)
                service_display = config.get_service_name_display()
                config.delete()
                messages.info(request, f'{service_display} API配置已删除')
                
            except APIConfig.DoesNotExist:
                messages.error(request, 'API配置不存在')
        
        return redirect('api_config')
    
    # 刷新配置列表
    configs = APIConfig.objects.all().order_by('service_name')
    
    context = {
        'configs': configs,
        'service_choices': APIConfig.SERVICE_CHOICES,
    }
    return render(request, 'core/api_config.html', context)

def test_api_connection(service_name, api_key):
    """测试API连接"""
    try:
        # 验证API密钥格式
        if not api_key or not api_key.strip() or len(api_key.strip()) < 10:
            print(f"API密钥格式无效: {service_name}")
            return False
            
        api_key = api_key.strip()
        
        if service_name == 'deepseek':
            url = 'https://api.deepseek.com/v1/chat/completions'
            headers = {
                'Content-Type': 'application/json',
                'Authorization': f'Bearer {api_key}'
            }
            payload = {
                'model': 'deepseek-chat',
                'messages': [{'role': 'user', 'content': 'Hello'}],
                'max_tokens': 10
            }
        elif service_name == 'kimi':
            url = 'https://api.moonshot.cn/v1/chat/completions'
            headers = {
                'Content-Type': 'application/json',
                'Authorization': f'Bearer {api_key}'
            }
            payload = {
                'model': 'moonshot-v1-8k',
                'messages': [{'role': 'user', 'content': 'Hello'}],
                'max_tokens': 10
            }
        else:
            print(f"不支持的服务类型: {service_name}")
            return False
        
        response = requests.post(url, headers=headers, json=payload, timeout=10)
        if response.status_code != 200:
            print(f"API测试失败 - 状态码: {response.status_code}, 响应: {response.text[:200]}")
        return response.status_code == 200
        
    except Exception as e:
        print(f"API测试异常: {service_name} - {e}")
        return False

def init_system_view(request):
    """系统初始化视图"""
    from django.core.management import execute_from_command_line
    from django.db import connection
    import sys
    
    init_results = []
    
    try:
        # 1. 检查数据库连接
        with connection.cursor() as cursor:
            cursor.execute("SELECT 1")
        init_results.append({"step": "数据库连接", "status": "成功", "message": "数据库连接正常"})
    except Exception as e:
        init_results.append({"step": "数据库连接", "status": "失败", "message": f"数据库连接失败: {e}"})
    
    try:
        # 2. 确保projects目录存在
        projects_root = str(settings.PROJECTS_ROOT)
        os.makedirs(projects_root, exist_ok=True)
        init_results.append({"step": "项目目录", "status": "成功", "message": f"项目根目录已创建: {projects_root}"})
    except Exception as e:
        init_results.append({"step": "项目目录", "status": "失败", "message": f"创建项目目录失败: {e}"})
    
    try:
        # 3. 检查并创建所有项目的目录结构
        projects = Project.objects.all()
        created_count = 0
        for project in projects:
            if not project.directory_path or not os.path.exists(project.directory_path):
                create_project_directory_structure(project)
                created_count += 1
        
        if created_count > 0:
            init_results.append({"step": "项目目录结构", "status": "成功", "message": f"已为 {created_count} 个项目创建目录结构"})
        else:
            init_results.append({"step": "项目目录结构", "status": "成功", "message": "所有项目目录结构已存在"})
    except Exception as e:
        init_results.append({"step": "项目目录结构", "status": "失败", "message": f"创建项目目录结构失败: {e}"})
    
    try:
        # 4. 检查静态文件
        static_root = os.path.join(settings.BASE_DIR, 'staticfiles')
        if os.path.exists(static_root):
            init_results.append({"step": "静态文件", "status": "成功", "message": "静态文件目录存在"})
        else:
            init_results.append({"step": "静态文件", "status": "警告", "message": "静态文件目录不存在，请运行 collectstatic"})
    except Exception as e:
        init_results.append({"step": "静态文件", "status": "失败", "message": f"检查静态文件失败: {e}"})
    
    # 5. 系统信息
    try:
        total_projects = Project.objects.count()
        init_results.append({"step": "系统状态", "status": "信息", "message": f"当前系统中共有 {total_projects} 个项目"})
    except Exception as e:
        init_results.append({"step": "系统状态", "status": "失败", "message": f"获取系统状态失败: {e}"})
    
    context = {
        'init_results': init_results,
        'success_count': len([r for r in init_results if r['status'] == '成功']),
        'error_count': len([r for r in init_results if r['status'] == '失败']),
        'warning_count': len([r for r in init_results if r['status'] == '警告']),
    }
    
    return render(request, 'core/init_system.html', context)

def test_upload_view(request):
    """测试文件上传页面"""
    return render(request, 'core/test_upload.html')

def get_network_config_api(request):
    """API: 获取网络共享配置"""
    config = get_network_config()
    return JsonResponse(config)

def get_network_config():
    """获取网络共享配置"""
    config_file = Path(settings.BASE_DIR) / 'network_config.json'
    default_config = {
        'network_share_path': '',  # 网络共享路径，如 \\192.168.1.100\projects
        'enable_network_share': False,  # 是否启用网络共享路径
        'enable_web_file_trial': True,  # 是否启用 Web 文件管理试用入口
    }
    
    if config_file.exists():
        try:
            with open(config_file, 'r', encoding='utf-8') as f:
                config = json.load(f)
                # 合并默认配置，确保新字段存在
                return {**default_config, **config}
        except Exception:
            pass
    return default_config

def save_network_config(config):
    """保存网络共享配置"""
    config_file = Path(settings.BASE_DIR) / 'network_config.json'
    with open(config_file, 'w', encoding='utf-8') as f:
        json.dump(config, f, ensure_ascii=False, indent=2)

def settings_view(request):
    """系统设置页面"""
    network_config = get_network_config()
    
    if request.method == 'POST':
        action = request.POST.get('action', 'save_path')
        
        if action == 'save_network':
            # 保存网络共享配置
            network_share_path = request.POST.get('network_share_path', '').strip()
            enable_network_share = request.POST.get('enable_network_share') == 'on'
            enable_web_file_trial = request.POST.get('enable_web_file_trial') == 'on'
            
            network_config['network_share_path'] = network_share_path
            network_config['enable_network_share'] = enable_network_share
            network_config['enable_web_file_trial'] = enable_web_file_trial
            save_network_config(network_config)
            
            if enable_network_share and network_share_path:
                messages.success(request, f'网络共享路径已配置为: {network_share_path}')
            else:
                messages.info(request, '已禁用网络共享路径，将使用本地路径')
            if enable_web_file_trial:
                messages.info(request, 'Web 文件管理试用入口已启用')
            else:
                messages.info(request, 'Web 文件管理试用入口已关闭')
        elif action == 'save_path':
            # 保存项目路径
            projects_root = request.POST.get('projects_root', '').strip()
            migrate_existing = request.POST.get('migrate_existing') == 'on'
            
            if projects_root:
                # 验证路径格式
                try:
                    path_obj = Path(projects_root)
                    # 检查路径是否为绝对路径
                    if not path_obj.is_absolute():
                        messages.error(request, '请输入绝对路径（完整路径）')
                    else:
                        # 更新settings.py文件
                        settings_file = Path(settings.BASE_DIR) / 'project_manager' / 'settings.py'
                        
                        with open(settings_file, 'r', encoding='utf-8') as f:
                            content = f.read()
                        
                        # 替换PROJECTS_ROOT配置
                        import re
                        pattern = r"PROJECTS_ROOT = .*"
                        if os.name == 'nt':  # Windows
                            replacement = f"PROJECTS_ROOT = Path(r'{projects_root}')"
                        else:  # Linux/Mac
                            replacement = f"PROJECTS_ROOT = Path('{projects_root}')"
                        
                        new_content = re.sub(pattern, replacement, content)
                        
                        with open(settings_file, 'w', encoding='utf-8') as f:
                            f.write(new_content)
                        
                        settings.PROJECTS_ROOT = Path(projects_root)
                        os.makedirs(projects_root, exist_ok=True)
                        
                        moved_count = 0
                        updated_count = 0
                        for project in Project.objects.all():
                            new_folder_name = _get_project_folder_name(project)
                            new_path = os.path.join(projects_root, new_folder_name)
                            
                            if migrate_existing and project.directory_path and os.path.exists(project.directory_path):
                                if _normalize_path(project.directory_path) != _normalize_path(new_path):
                                    moved, _ = _move_path_with_backup(project.directory_path, new_path)
                                    if moved:
                                        moved_count += 1
                            
                            if project.directory_path != new_path:
                                project.directory_path = new_path
                                project.save(update_fields=['directory_path'])
                                updated_count += 1
                        
                        if migrate_existing:
                            messages.success(request, f'项目路径已更新为: {projects_root}。已更新 {updated_count} 个项目目录路径，迁移 {moved_count} 个目录。请重启服务器以使更改生效。')
                        else:
                            messages.success(request, f'项目路径已更新为: {projects_root}。已更新 {updated_count} 个项目目录路径。请重启服务器以使更改生效。')
                        
                except Exception as e:
                    messages.error(request, f'路径格式错误: {str(e)}')
            else:
                messages.error(request, '请输入有效的路径')
        elif action == 'update_project_path':
            project_id = request.POST.get('project_id', '').strip()
            project_path = request.POST.get('project_path', '').strip()
            move_project_files = request.POST.get('move_project_files') == 'on'
            
            if not project_id or not project_path:
                messages.error(request, '请输入项目编号和目标路径')
            else:
                try:
                    path_obj = Path(project_path)
                    if not path_obj.is_absolute():
                        messages.error(request, '请输入绝对路径（完整路径）')
                    else:
                        project = get_object_or_404(Project, project_id=project_id)
                        if move_project_files and project.directory_path and os.path.exists(project.directory_path):
                            _move_path_with_backup(project.directory_path, project_path)
                        os.makedirs(project_path, exist_ok=True)
                        project.directory_path = project_path
                        project.save(update_fields=['directory_path'])
                        messages.success(request, f'项目 {project.project_id} 路径已更新为: {project_path}')
                except Exception as e:
                    messages.error(request, f'更新项目路径失败: {str(e)}')
    
    # 获取当前配置的项目路径
    current_projects_root = str(settings.PROJECTS_ROOT)
    network_config = get_network_config()  # 重新获取最新配置
    projects = Project.objects.all().order_by('project_id')
    
    context = {
        'current_projects_root': current_projects_root,
        'network_share_path': network_config.get('network_share_path', ''),
        'enable_network_share': network_config.get('enable_network_share', False),
        'enable_web_file_trial': network_config.get('enable_web_file_trial', True),
        'projects': projects,
    }
    
    return render(request, 'core/settings.html', context)
